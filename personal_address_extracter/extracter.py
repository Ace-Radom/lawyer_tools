import openpyxl
import os
import warnings

os.environ["PADDLE_OCR_BASE_DIR"] = "D:\paddleOCR"
warnings.filterwarnings("ignore", category=UserWarning)

from datetime import datetime
from openpyxl.styles import Alignment
from paddleocr import PaddleOCR
from rich import print as rprint
from rich.panel import Panel

ocr_worker = PaddleOCR(use_angle_cls=True, show_log=False, lang="ch")

TARGET_IMAGE_DIRNAME = "images"

TAG_NAME = "姓名"
TAG_SEX = "性别"
TAG_ID = "公民身份证号"
TAG_BIRTH = "出生日期"
TAG_ADDR = "户籍地址"

SAME_LINE_Y_MAX_PIXEL_OFFSET = 50
SAME_TAG_DATA_X_MAX_PIXEL_OFFSET = 20

class coord:
    def __init__(self) -> None:
        self.x = -1
        self.y = -1
        return

    def __init__(self, x: float, y: float) -> None:
        self.x = x
        self.y = y
        return

    def empty(self) -> bool:
        return self.x == -1 and self.y == -1
    
# extracted persional data structure in one file 
class personal_data:
    def __init__(self) -> None:
        self.name = ""
        self.sex = ""
        self.id = ""
        self.birth = ""
        self.addr = ""
        return
    
    def empty(self) -> None:
        return len(self.name) == 0 and len(self.sex) == 0 and len(self.id) == 0 and len(self.birth) == 0 and len(self.addr) == 0

def perform_ocr(image_path: str) -> list:
    if not os.path.isfile(image_path):
        rprint(f"[red]Error[/]: image `{image_path}` doesn't exist.")
        return []
    elif os.path.splitext(image_path)[1] not in [".jpg", ".jpeg", ".png"]:
        rprint(f"[red]Error[/]: image `{image_path}` has a not supported extension.")
        return []
    
    try:
        result = ocr_worker.ocr(image_path)
    except Exception as ex:
        rprint(f"[red]Error[/]: failed to perform OCR on image `{image_path}`.\n", ex)

    return result[0]

def check_ocr_result(result: list) -> bool:
    if len(result) < 1:
        rprint("[red]Error[/]: result length less than 1.")
        return False
    else:
        tag_name_found = False
        tag_sex_found = False
        tag_id_found = False
        tag_birth_found = False
        tag_addr_found = False

        for line in result:
            recognized_text = line[1][0]
            if TAG_NAME in recognized_text: tag_name_found = True
            elif TAG_SEX in recognized_text: tag_sex_found = True
            elif TAG_ID in recognized_text: tag_id_found = True
            elif TAG_BIRTH in recognized_text: tag_birth_found = True
            elif TAG_ADDR in recognized_text: tag_addr_found = True
            else: continue

        if not (tag_name_found and tag_sex_found and tag_id_found and tag_birth_found and tag_addr_found):
            rprint("[red]Error[/]: not all necessary tag found.")
            return False
    return True

def get_uo_coord(coords: list[list[float]]) -> coord:
    # (0,0) means under-left corner of the picture (by default)
    p0_min = 65536
    p0_max = -1
    p1_min = 65536
    p1_max = -1
    for co in coords:
        if co[0] > p0_max: p0_max = co[0]
        if co[0] < p0_min: p0_min = co[0]
        if co[1] > p1_max: p1_max = co[1]
        if co[1] < p1_min: p1_min = co[1]
    ordered_coords = []
    if (p0_max - p0_min) >= (p1_max - p1_min):
        # (x,y)
        for co in coords:
            ordered_coords.append(coord(co[0], co[1]))
    else:
        # (y,x)
        for co in coords:
            ordered_coords.append(coord(co[1], co[0]))
    # x-offset should always bigger than y-offset

    return ordered_coords[0]
    # coord order: lu -> lo -> ro -> ru
    # TODO: orders of coords may be changed in different models, a reorder may be necessary

def find_data_with_tag(result, tag: str, allow_multiline: bool) -> tuple[str, float]:
    tag_coord_uo: coord = {}
    for line in result:
        recognized_text = line[1][0]
        if recognized_text == tag:
            tag_coord_uo = get_uo_coord(line[0])
            break
    if tag_coord_uo.empty():
        rprint(f"[red]Error[/]: tag `{tag}` not found in OCR result.")
        return ("", -1)
    
    closest_line = []
    possible_lines = []
    
    for line in result:
        recognized_text = line[1][0]
        if recognized_text == tag:
            continue
        # skip tag itself
        
        this_coord_uo = get_uo_coord(line[0])
        if abs(this_coord_uo.y - tag_coord_uo.y) < SAME_LINE_Y_MAX_PIXEL_OFFSET and this_coord_uo.x > tag_coord_uo.x:
            if allow_multiline:
                possible_lines.append(line)
            elif len(closest_line) != 0:
                closest_line_uo = get_uo_coord(closest_line[0])
                if closest_line_uo.x <= this_coord_uo.x or abs(closest_line_uo.y - tag_coord_uo.y) < abs(this_coord_uo.y - tag_coord_uo.y):
                    continue
                # closest line until now has a closer position to tag or a smaller y-offset, skip this line
            # closest line not empty, judge which one is closer
            closest_line = line
    
    if allow_multiline:
        if len(possible_lines) == 0:
            rprint(f"[red]Error[/]: no data assigned to tag `{tag}` was found.")
            return ("", -1)
        for i in range(0, len(possible_lines)):
            for j in range(0 , i):
                if get_uo_coord(possible_lines[i][0]).y < get_uo_coord(possible_lines[j][0]).y:
                    temp = possible_lines[i]
                    possible_lines[i] = possible_lines[j]
                    possible_lines[j] = temp
        if len(possible_lines) > 2:
            data = possible_lines[-2][1][0] + possible_lines[-1][1][0]
            acc = (possible_lines[-2][1][1] + possible_lines[-1][1][1]) / 2
            return (data, acc)
        elif len(possible_lines) < 2:
            return (possible_lines[0][1][0], possible_lines[0][1][1])
        else:
            if abs(get_uo_coord(possible_lines[0][0]).x - get_uo_coord(possible_lines[1][0]).x) > SAME_TAG_DATA_X_MAX_PIXEL_OFFSET:
                return(possible_lines[-1][1][0], possible_lines[-1][1][1])
            # two lines were found but one doesn't belong to this tag
            data = possible_lines[0][1][0] + possible_lines[1][1][0]
            acc = (possible_lines[0][1][1] + possible_lines[1][1][1]) / 2
            return (data, acc)
    else:
        if len(closest_line) == 0:
            rprint(f"[red]Error[/]: no data assigned to tag `{tag}` was found.")
            return ("", -1)
        return (closest_line[1][0], closest_line[1][1])

def extract_picture_personal_data(image_path: str) -> personal_data:
    result = perform_ocr(image_path)
    if not check_ocr_result(result):
        return personal_data()
    # illegal ocr result

    data = personal_data()

    data.name, acc = find_data_with_tag(result, TAG_NAME, allow_multiline=False)
    if acc == -1:
        return personal_data()
    elif acc < 0.95:
        rprint("[yellow]Warning[/]: low accuracy expected on `name`.")

    data.sex, acc = find_data_with_tag(result, TAG_SEX, allow_multiline=False)
    if acc == -1:
        return personal_data()
    elif acc < 0.95:
        rprint("[yellow]Warning[/]: low accuracy expected on `sex`.")

    data.id, acc = find_data_with_tag(result, TAG_ID, allow_multiline=False)
    if acc == -1:
        return personal_data()
    elif acc < 0.95:
        rprint("[yellow]Warning[/]: low accuracy expected on `id`.")

    data.birth, acc = find_data_with_tag(result, TAG_BIRTH, allow_multiline=False)
    if acc == -1:
        return personal_data()
    elif acc < 0.95:
        rprint("[yellow]Warning[/]: low accuracy expected on `birth`.")

    data.addr, acc = find_data_with_tag(result, TAG_ADDR, allow_multiline=True)
    if acc == -1:
        return personal_data()
    elif acc < 0.95:
        rprint("[yellow]Warning[/]: low accuracy expected on `addr`.")

    return data

def main():
    os.chdir(os.path.dirname(__file__))
    if not os.path.isdir(TARGET_IMAGE_DIRNAME):
        rprint(f"[red]Error[/]: target image directory `{TARGET_IMAGE_DIRNAME}` doesn't exist.")
        return 1
    personal_datas = []
    for _, _, files in os.walk(TARGET_IMAGE_DIRNAME):
        rprint(f"Total [cyan]{len(files)}[/] files found.")
        for file in files:
            rprint(f"Extracting personal datas from file `[cyan]{file}[/]`...")
            data = extract_picture_personal_data(os.path.join(TARGET_IMAGE_DIRNAME, file))
            if data.empty():
                personal_datas.append({
                    "filename": file,
                    "data": personal_data()
                })
                continue
            # OCR failed, push an empty data into list
            panel_text = f" [yellow]•[/] [green]{TAG_NAME}[/]: {data.name}\n"
            panel_text += f" [yellow]•[/] [green]{TAG_SEX}[/]: {data.sex}\n"
            panel_text += f" [yellow]•[/] [green]{TAG_ID}[/]: {data.id}\n"
            panel_text += f" [yellow]•[/] [green]{TAG_BIRTH}[/]: {data.birth}\n"
            panel_text += f" [yellow]•[/] [green]{TAG_ADDR}[/]: {data.addr}"
            panel = Panel(
                panel_text,
                title=f"Datas from `[cyan]{file}[/]`"
            )
            rprint(panel)
            personal_datas.append({
                "filename": file,
                "data": data
            })

    now_str = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f"户籍数据汇总"
    # sheet title
    sheet["B1"] = "源文件名"
    sheet["C1"] = TAG_NAME
    sheet["D1"] = TAG_SEX
    sheet["E1"] = TAG_ID
    sheet["F1"] = TAG_BIRTH
    sheet["G1"] = TAG_ADDR
    # header
    center_alignment = Alignment(horizontal='center', vertical='center')
    sheet["B1"].alignment = center_alignment
    sheet["C1"].alignment = center_alignment
    sheet["D1"].alignment = center_alignment
    sheet["E1"].alignment = center_alignment
    sheet["F1"].alignment = center_alignment
    sheet["G1"].alignment = center_alignment
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 15
    sheet.column_dimensions["G"].width = 100
    # style

    for index, this_data in enumerate(personal_datas):
        filename = this_data["filename"]
        data = this_data["data"]
        row_index = index + 2
        sheet.cell(row_index, 1, index + 1)
        sheet.cell(row_index, 2, filename)
        if data.empty():
            continue
        # failed result
        sheet.cell(row_index, 3, data.name)
        sheet.cell(row_index, 4, data.sex)
        sheet.cell(row_index, 5, data.id)
        sheet.cell(row_index, 6, data.birth)
        sheet.cell(row_index, 7, data.addr)

    wb.save(f"户籍数据汇总.{now_str}.xlsx")

if __name__ == "__main__":
    main()
